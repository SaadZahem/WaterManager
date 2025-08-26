from json import load
from typing import NamedTuple
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook

path = "G:\\New folder\\Egyptian Development Bank\\data\\"


class UserData(NamedTuple):
    name: str
    gender: str
    nid: str = ""
    country: str = ""
    gov: str = ""
    tele: str = ""
    email: str = ""
    password: str = ""


class UserDataStorage(object):
    filename = path + "users.xlsx"
    file = None

    @classmethod
    def init(cls):
        try:
            cls.file = load_workbook(filename=cls.filename, read_only=False)
        except FileNotFoundError:
            cls.file = Workbook(write_only=False)
            sheet = cls.file.active
            sheet["A1"] = "No."
            sheet["B1"] = "Name"
            sheet["C1"] = "Gender"
            sheet["D1"] = "National ID"
            sheet["E1"] = "Country"
            sheet["F1"] = "Governorate"
            sheet["G1"] = "Telephone number"
            sheet["H1"] = "Email"
            sheet["I1"] = "Password"
            cls._save()

    @classmethod
    def _save(cls):
        cls.file.save(filename=cls.filename)

    @classmethod
    def patch(cls, user: UserData):
        ws = cls.file.active
        for i, cell in enumerate(ws["B"], start=1):
            if i == 1:
                continue
            if cell.value == user.name:
                ws["C%i" % i] = user.gender or ws["C%i" % i].value
                ws["D%i" % i] = user.nid or ws["D%i" % i].value
                ws["E%i" % i] = user.country or ws["E%i" % i].value
                ws["F%i" % i] = user.gov or ws["F%i" % i].value
                ws["G%i" % i] = user.tele or ws["G%i" % i].value
                ws["H%i" % i] = user.email or ws["H%i" % i].value
                ws["I%i" % i] = user.password or ws["I%i" % i].value
                break
        else:
            i += 1
            ws["B%i" % i] = user.name
            ws["C%i" % i] = user.gender
            ws["D%i" % i] = user.nid
            ws["E%i" % i] = user.country
            ws["F%i" % i] = user.gov
            ws["G%i" % i] = user.tele
            ws["H%i" % i] = user.email
            ws["I%i" % i] = user.password
        cls._save()


if __name__ == "__main__":
    UserDataStorage.init()
    UserDataStorage.patch(UserData("Nader2", "male", email="sample@gmail.com"))
